from openpyxl import load_workbook  # 读取excel的库
from matplotlib.colors import LinearSegmentedColormap  # 色带
from matplotlib import pyplot as plt  # 画图框架
from seaborn import heatmap  # 热力图

# 输入文件名，导出excel内容和其中的最大最小值
def eyeDiagram(fileName='./data/eyedata.xlsx'):
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 用于matplotlib显示中文
    plt.rcParams['axes.unicode_minus'] = False  # 用于matplotlib显示中文
    bk = load_workbook(fileName)
    sheet = bk.active
    minrow = sheet.min_row
    maxrow = sheet.max_row
    mincol = sheet.min_column
    maxcol = sheet.max_column
    data_xlsx = []
    maxNum = 0
    minNum = 0
    for i in range(minrow+1, maxrow + 1):
        coldata = []
        for j in range(mincol+1, maxcol + 1):
            cell = sheet.cell(i, j).value
            if cell != None:
                if int(cell) > maxNum:
                    maxNum = cell
                if int(cell) < minNum:
                    minNum = cell
                coldata.append(cell)
        data_xlsx.append(coldata)
    bk.close()
    plt.figure(dpi=100)  # 设置窗口大小
    # 定义渐变颜色条（蓝色-红色）
    my_colormap = LinearSegmentedColormap.from_list("",
                                                    ["blue", "Deepskyblue", 'yellow', 'orange', 'red', "Darkred"],
                                                    N=14)  # N=14表示将色条对应14个色段 颜色对照表网站https://www.qianbo.org/Tool/Rgba/
    heatmap(data=data_xlsx, vmax=maxNum, vmin=minNum,
                cmap=my_colormap,annot=False)  # vmax输入色号对应最大值，vmin输入色号对应最小值，annot是否显示数值，linewidths设置单元格边框宽度
    plt.title("眼图示例")
    plt.ylabel('y_label')
    plt.xlabel('x_label')
    plt.show()